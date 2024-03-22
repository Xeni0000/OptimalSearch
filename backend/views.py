from django.db.models import Count, Q
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from backend.common.arch.forms import WriteToDBForm, PaginationForm, SearchForm, OrderForm
from backend.common.decorators import require_POST, require_GET
from backend.common.http import ErrorResponse, SuccessResponse
from backend.models import Subsystem, Folder, Task, Segment, Attachment, Role, Template
from backend.serv.WriteOnDBServ import WriteOnDBServ


@require_POST
async def write_on_db(request):
    """Функция бля построения зависимостей и записи данных в БД"""

    form = WriteToDBForm.from_POST(request)

    # считываем при помощи библиотеки openpyxl данные из файлов
    try:
        subsystems_wb = load_workbook(form.path_to_tasks)
    except InvalidFileException as ex:
        return ErrorResponse({'path_to_tasks': str(ex)})
    except FileNotFoundError as ex:
        return ErrorResponse({'path_to_tasks': str(ex)})

    try:
        templates_wb = load_workbook(form.path_to_templates)
    except InvalidFileException as ex:
        return ErrorResponse({'path_to_templates': str(ex)})
    except FileNotFoundError as ex:
        return ErrorResponse({'path_to_templates': str(ex)})

    try:
        roles_wb = load_workbook(form.path_to_roles)
    except InvalidFileException as ex:
        return ErrorResponse({'path_to_roles': str(ex)})
    except FileNotFoundError as ex:
        return ErrorResponse({'path_to_roles': str(ex)})

    try:
        segments_wb = load_workbook(form.path_to_roles_list_wb)
    except InvalidFileException as ex:
        return ErrorResponse({'path_to_roles_list_wb': str(ex)})
    except FileNotFoundError as ex:
        return ErrorResponse({'path_to_roles_list_wb': str(ex)})

    # region Templates

    templates_ws = templates_wb.active
    templates_list = [t for t in templates_ws.values]
    templates = await WriteOnDBServ.write_templates(templates_list[1:])

    # endregion

    # region Segments

    segments_ws = segments_wb.active
    segments_list = segments_ws.values
    segments, roles = await WriteOnDBServ.write_segments(segments_list)

    # endregion

    # region Tasks
    subsystems_ws = subsystems_wb.active
    subsystems_list = subsystems_ws.values
    subsystems, templates, roles = await WriteOnDBServ.write_subsystems(subsystems_list, templates, roles)

    # endregion

    # region Roles

    roles_ws = roles_wb.active
    roles_list = [r for r in roles_ws.values]
    templates, roles, subsystems, tasks = await WriteOnDBServ.write_attachments(roles_list[1:], templates, roles,
                                                                                subsystems, segments)

    # endregion
    not_added_roles: list[str] = []
    not_added_tasks: list[str] = []

    for r in roles.values():
        r_templates = [v.template for k, v in templates.items() if k in r.templates]

        if not r.role:
            not_added_roles.append(r.name)
            continue

        await r.role.templates.aadd(*r_templates)

    for t in tasks.values():
        t_roles = [v.role for k, v in roles.items() if k in t.roles]

        if not t.task:
            not_added_tasks.append(t.name)
            continue

        await t.task.roles.aadd(*t_roles)

    return SuccessResponse({
        'not_added_roles': not_added_roles,
        'not_added_tasks': not_added_tasks,
    })


@require_GET
async def role_list(request):
    pag_form = PaginationForm.from_GET(request)
    search_form = SearchForm.from_GET(request)
    order_form = OrderForm.from_GET(request)

    roles = Role.objects.select_related('attachment', 'attachment__segment').filter()

    if search_form.search_text:
        roles = roles.filter(Q(name__icontains=search_form.search_text) |
                             Q(attachment__name__icontains=search_form.search_text) |
                             Q(attachment__segment__name__icontains=search_form.search_text))

    if order_form.order_by:
        roles = roles.order_by(order_form.ordering_str)
    else:
        roles = roles.order_by('name')

    return SuccessResponse({'roles': [
        await r.adict(attachment_in=True) async for r in roles[pag_form.fr: pag_form.to]
    ],
        'count': await roles.acount()})


@require_GET
async def task_list(request):
    pag_form = PaginationForm.from_GET(request)
    search_form = SearchForm.from_GET(request)
    order_form = OrderForm.from_GET(request)

    tasks = Task.objects.select_related('folder', 'folder__subsystem').filter()

    if search_form.search_text:
        tasks = tasks.filter(Q(name__icontains=search_form.search_text) |
                             Q(folder__name__icontains=search_form.search_text) |
                             Q(folder__subsystem__name__icontains=search_form.search_text))

    if order_form.order_by:
        tasks = tasks.order_by(order_form.ordering_str)
    else:
        tasks = tasks.order_by('name')

    return SuccessResponse({'tasks': [
        await t.adict(folder_in=True) async for t in tasks[pag_form.fr: pag_form.to]
    ],
        'count': await tasks.acount()})


@require_POST
async def find_tasks_duplicate_dependence(request):
    """Функция для поиска повторяющихся задачей в БД"""

    # Получаем из БД роли и добавляем к каждой из них поле с указанием количества связанных задач
    roles_with_task_count = Role.objects.annotate(task_count=Count('task'))

    # фильтруем роли по количеству задач больше одной
    roles_with_multiple_tasks = roles_with_task_count.filter(task_count__gt=1)

    # Ищем задачи, роли которых мы отфильтровали выше
    duplicated_tasks = Task.objects.filter(roles__in=roles_with_multiple_tasks)

    # возвращаем список задач и связанных ролей
    return SuccessResponse({
        'tasks': [{
            'task': t.dict(),
            'roles': [r.dict() async for r in t.roles.all()]
        } async for t in duplicated_tasks],
        'count': await duplicated_tasks.acount()
    })
